import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from io import BytesIO

# تنظیمات اولیه صفحه استریم‌لیت
st.set_page_config(page_title="مدیریت و فیلتر ماموریت‌های پرسنلی", layout="wide")

st.title("📊 سامانه بررسی و نشانه‌گذاری داده‌های پرسنلی")
st.write("فایل اکسل خود را آپلود کنید، پرسنل یکتا را بررسی کنید و ردیف‌های مورد نظر را با خط قرمز در فایل خروجی متمایز کنید.")

# ۱. دریافت فایل از کاربر
uploaded_file = st.file_uploader("لطفاً فایل اکسل (xlsx) خود را انتخاب و آپلود کنید:", type=["xlsx"])

if uploaded_file is not None:
    try:
        # خواندن داده‌ها با پانداس برای پردازش و نمایش سریع
        # خواندن همه ستون‌ها به صورت رشته برای جلوگیری از حذف صفرهای اول کد پرسنلی
        df = pd.read_excel(uploaded_file, dtype={'کد پرسنلی': str})
        
        # بررسی وجود ستون‌های حیاتی در فایل
        required_columns = ['کد پرسنلی', 'نام و نام خانوادگی', 'تاریخ شروع', 'ساعت شروع', 'ساعت پایان']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            st.error(f"خطا: ستون‌های زیر در فایل آپلود شده یافت نشدند: {missing_columns}")
        else:
            # ایجاد دیتافریم از افراد یکتا برای ساخت چک‌باکس‌ها
            # ترکیب کد پرسنلی و نام برای اطمینان از یکتا بودن کلید
            df_unique = df[['کد پرسنلی', 'نام و نام خانوادگی']].drop_duplicates().reset_index(drop=True)
            
            st.subheader("👥 لیست پرسنل یکتا و مدیریت وضعیت")
            st.info("با کلیک روی نام هر فرد، جزئیات رکوردها (تاریخ و ساعت) نمایش داده می‌شود. تیک زدن چک‌باکس باعث کشیده شدن خط قرمز روی تمام ردیف‌های آن شخص در فایل خروجی نهایی می‌شود.")
            
            # ذخیره وضعیت تیک خوردن چک‌باکس‌ها در session_state استریم‌لیت
            if 'selected_personnel' not in st.session_state:
                st.session_state.selected_personnel = set()
                
            # نمایش داده‌ها به صورت ساختار یافته
            for index, row in df_unique.iterrows():
                emp_id = row['کد پرسنلی']
                emp_name = row['نام و نام خانوادگی']
                unique_key = f"{emp_id} - {emp_name}"
                
                # ایجاد دو ستون: یکی برای چک‌باکس و یکی برای دکمه جزئیات
                col1, col2 = st.columns([1, 10])
                
                with col1:
                    # بررسی اینکه آیا قبلاً تیک خورده یا خیر
                    is_checked = unique_key in st.session_state.selected_personnel
                    if st.checkbox("", key=f"chk_{emp_id}", value=is_checked):
                        st.session_state.selected_personnel.add(unique_key)
                    else:
                        st.session_state.selected_personnel.discard(unique_key)
                        
                with col2:
                    # دکمه بازشوندگی (Expander) برای نمایش ردیف‌های مربوط به این شخص
                    with st.expander(f"🆔 کد: {emp_id} | 👤 نام: {emp_name}"):
                        # فیلتر کردن ردیف‌های اصلی مربوط به این شخص
                        person_details = df[df['کد پرسنلی'] == emp_id][['تاریخ شروع', 'ساعت شروع', 'ساعت پایان']]
                        st.dataframe(person_details, use_container_width=True)
            
            st.markdown("---")
            st.subheader("📥 دانلود فایل خروجی")
            
            # دکمه پردازش و ساخت فایل خروجی اکسل با استایل خط قرمز
            if st.button("🔄 پردازش و اعمال خط تیره قرمز روی ردیف‌های انتخابی"):
                if not st.session_state.selected_personnel:
                    st.warning("هیچ پرسنلی انتخاب نشده است. فایل خروجی بدون تغییر خواهد بود.")
                
                # بازنشانی نشانگر فایل آپلود شده برای خواندن مجدد توسط openpyxl
                uploaded_file.seek(0)
                wb = openpyxl.load_workbook(uploaded_file)
                ws = wb.active
                
                # پیدا کردن اندیس ستون‌های "کد پرسنلی" و "نام و نام خانوادگی" در اکسل (داده‌ها معمولاً از ردیف ۱ شروع می‌شوند)
                header_row = [cell.value for cell in ws[1]]
                
                try:
                    id_col_idx = header_row.index('کد پرسنلی') + 1
                    name_col_idx = header_row.index('نام و نام خانوادگی') + 1
                    
                    # تعریف استایل فونت: رنگ قرمز و دارای خط خوردگی (Strikethrough)
                    red_strikethrough_font = Font(strikethrough=True, color="FF0000")
                    
                    # پیمایش ردیف‌های اکسل از ردیف دوم به بعد (ردیف اول هدر است)
                    marked_rows_count = 0
                    for row_idx in range(2, ws.max_row + 1):
                        cell_id = str(ws.cell(row=row_idx, column=id_col_idx).value).strip()
                        cell_name = str(ws.cell(row=row_idx, column=name_col_idx).value).strip()
                        
                        # بررسی مطابقت ردیف فعلی با لیست پرسنل تیک‌خورده
                        current_row_key = f"{cell_id} - {cell_name}"
                        
                        if current_row_key in st.session_state.selected_personnel:
                            # اعمال فونت قرمز خط‌خورده روی تک تک سلول‌های این ردیف
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = red_strikethrough_font
                            marked_rows_count += 1
                    
                    # ذخیره فایل در حافظه موقت (BytesIO) برای دانلود مستقیم در استریم‌لیت
                    output_stream = BytesIO()
                    wb.save(output_stream)
                    output_stream.seek(0)
                    
                    st.success(f"عملیات با موفقیت انجام شد! تعداد {marked_rows_count} ردیف در اکسل خط‌خوردگی قرمز دریافت کردند.")
                    
                    # ایجاد دکمه دانلود نهایی در مرورگر کاربر
                    st.download_button(
                        label="📥 دانلود فایل اکسل اصلاح شده",
                        data=output_stream,
                        file_name="processed_missions.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except ValueError:
                    st.error("خطا در ساختار هدر فایل اکسل. مطمئن شوید ستون‌های 'کد پرسنلی' و 'نام و نام خانوادگی' دقیقاً با همین نام‌ها وجود دارند.")
                    
    except Exception as e:
        st.error(f"خطا در خواندن یا پردازش فایل: {e}")
