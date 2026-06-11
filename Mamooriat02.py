import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import re
from datetime import datetime

# تنظیمات اولیه صفحه استریم‌لیت
st.set_page_config(page_title="مدیریت و فیلتر ماموریت‌های پرسنلی", layout="wide")

st.title("📊 سامانه هوشمند بررسی ماموریت و مغایرت‌گیری تردد پرسنل")
st.write("فایل ماموریت‌ها را آپلود کنید، ترددهای هر شخص را به صورت مجزا بررسی کنید و ردیف‌های مغایرت دار یا تایید شده را مشاهده کنید.")

# تابع کمکی برای پاک‌سازی و استخراج تاریخ شمسی از ستون 'روز' فایل تردد
def extract_shamsi_date(text):
    if pd.isna(text):
        return None
    text = str(text).strip()
    # پیدا کردن الگوی تاریخ شمسی مانند 1405/02/26
    match = re.search(r'\d{4}/\d{2}/\d{2}', text)
    if match:
        return match.group(0)
    return None

# تابع کمکی برای تبدیل رشته ساعت (HH:MM) به دقیقه جهت مقایسه ریاضی راحت‌تر
def time_to_minutes(time_str):
    if pd.isna(time_str) or not str(time_str).strip() or str(time_str).strip() == 'nan':
        return None
    try:
        time_str = str(time_str).strip()
        parts = time_str.split(':')
        if len(parts) >= 2:
            return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return None
    return None

# ۱. دریافت فایل اصلی ماموریت‌ها از کاربر
uploaded_file = st.file_uploader("۱. لطفاً فایل اصلی ماموریت‌ها (xlsx) را آپلود کنید:", type=["xlsx"], key="main_file")

if uploaded_file is not None:
    try:
        # خواندن داده‌های ماموریت با پانداس
        df = pd.read_excel(uploaded_file, dtype={'کد پرسنلی': str})
        
        # بررسی وجود ستون‌های حیاتی در فایل ماموریت‌ها
        required_columns = ['کد پرسنلی', 'نام و نام خانوادگی', 'تاریخ شروع', 'ساعت شروع', 'ساعت پایان']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            st.error(f"خطا: ستون‌های زیر در فایل ماموریت‌ها یافت نشدند: {missing_columns}")
        else:
            # ایجاد دیتافریم از افراد یکتا برای ساخت چک‌باکس‌ها
            df_unique = df[['کد پرسنلی', 'نام و نام خانوادگی']].drop_duplicates().reset_index(drop=True)
            
            st.subheader("👥 لیست پرسنل یکتا و مدیریت وضعیت")
            st.info("با باز کردن منوی هر شخص، علاوه بر رویت اطلاعات ماموریت، می‌توانید فایل اکسل تردد همان شخص را آپلود کرده و مغایرت‌گیری کنید.")
            
            # ذخیره وضعیت تیک خوردن چک‌باکس‌ها در session_state استریم‌لیت
            if 'selected_personnel' not in st.session_state:
                st.session_state.selected_personnel = set()
                
            # نمایش داده‌ها به صورت ساختار یافته برای هر فرد یکتا
            for index, row in df_unique.iterrows():
                emp_id = row['کد پرسنلی']
                emp_name = row['نام و نام خانوادگی']
                unique_key = f"{emp_id} - {emp_name}"
                
                # ایجاد دو ستون: یکی برای چک‌باکس حذف/نشانه‌گذاری و یکی برای دکمه جزئیات
                col1, col2 = st.columns([1, 20])
                
                with col1:
                    is_checked = unique_key in st.session_state.selected_personnel
                    if st.checkbox("", key=f"chk_{emp_id}", value=is_checked):
                        st.session_state.selected_personnel.add(unique_key)
                    else:
                        st.session_state.selected_personnel.discard(unique_key)
                        
                with col2:
                    # دکمه بازشوندگی (Expander) برای هر شخص
                    with st.expander(f"🆔 کد: {emp_id} | 👤 نام: {emp_name}"):
                        
                        # ساب‌کالوم‌ها برای چیدمان داخلی منوی بازشونده
                        exp_col1, exp_col2 = st.columns([1, 1])
                        
                        with exp_col1:
                            st.markdown("📋 **اطلاعات ماموریت‌ها در فایل اصلی:**")
                            # فیلتر کردن ردیف‌های اصلی مربوط به این شخص
                            person_details = df[df['کد پرسنلی'] == emp_id][['تاریخ شروع', 'ساعت شروع', 'ساعت پایان']].copy()
                            st.dataframe(person_details, use_container_width=True, hide_index=True)
                        
                        with exp_col2:
                            st.markdown("⏱️ **آپلوُد و مغایرت‌گیری اطلاعات تردد همین فرد:**")
                            # آپلودر فایل اختصاصی تردد برای این فرد یکتا
                            traffic_file = st.file_uploader(f"فایل اکسل تردد ({emp_name})", type=["xlsx"], key=f"traffic_{emp_id}")
                            
                            if traffic_file is not None:
                                try:
                                    # خواندن فایل تردد
                                    df_traffic = pd.read_excel(traffic_file)
                                    
                                    # بررسی وجود ستون‌های مورد نیاز در فایل تردد (روز، شروع، پایان)
                                    if 'روز' in df_traffic.columns and 'شروع' in df_traffic.columns and 'پایان' in df_traffic.columns:
                                        
                                        # استخراج تاریخ دقیق شمسی برای مقایسه
                                        df_traffic['تاریخ_شمسی'] = df_traffic['روز'].apply(extract_shamsi_date)
                                        
                                        # لیست برای ذخیره نتایج مقایسه
                                        comparison_results = []
                                        
                                        # بررسی تک تک ماموریت‌های این شخص
                                        for _, m_row in person_details.iterrows():
                                            m_date = str(m_row['تاریخ شروع']).strip()
                                            m_start_str = str(m_row['ساعت شروع']).strip()
                                            m_end_str = str(m_row['ساعت پایان']).strip()
                                            
                                            m_start_min = time_to_minutes(m_start_str)
                                            m_end_min = time_to_minutes(m_end_str)
                                            
                                            # پیدا کردن ردیف تردد متناظر با تاریخ ماموریت (فقط ردیف‌هایی که وضعیت حضور دارند و ساعت خالی نیست)
                                            matched_traffic = df_traffic[
                                                (df_traffic['تاریخ_شمسی'] == m_date) & 
                                                (df_traffic['شروع'].notna()) & 
                                                (df_traffic['پایان'].notna())
                                            ]
                                            
                                            status = "❌ عدم تطابق تاریخ / بدون تردد"
                                            color = "#ff4b4b" # قرمز استریم‌لیت
                                            t_start_str, t_end_str = "-", "-"
                                            
                                            if not matched_traffic.empty:
                                                # گرفتن اولین تردد معتبر یافت شده در آن روز
                                                t_row = matched_traffic.iloc[0]
                                                t_start_str = str(t_row['شروع']).strip()
                                                t_end_str = str(t_row['پایان']).strip()
                                                
                                                t_start_min = time_to_minutes(t_start_str)
                                                t_end_min = time_to_minutes(t_end_str)
                                                
                                                if m_start_min is not None and m_end_min is not None and t_start_min is not None and t_end_min is not None:
                                                    # منطق اصلی: آیا ماموریت کاملاً درون بازه تردد است؟
                                                    if t_start_min <= m_start_min and m_end_min <= t_end_min:
                                                        status = "✅ تایید (داخل بازه تردد)"
                                                        color = "#23c16a" # سبز استریم‌لیت
                                                    else:
                                                        status = "🚨 مغایرت (خارج از بازه تردد)"
                                                        color = "#ff4b4b" # قرمز
                                            
                                            comparison_results.append({
                                                "تاریخ": m_date,
                                                "شروع ماموریت": m_start_str,
                                                "پایان ماموریت": m_end_str,
                                                "شروع تردد": t_start_str,
                                                "پایان تردد": t_end_str,
                                                "وضعیت بررسی": status,
                                                "رنگ": color
                                            })
                                        
                                        # نمایش نتایج با فرمت‌دهی رنگی زیبا زنده در UI
                                        st.markdown("**نتایج آنالیز هوشمند مغایرت‌ها:**")
                                        for res in comparison_results:
                                            # استفاده از HTML سفارشی برای ایجاد باکس‌های رنگی سبز و قرمز بر اساس منطق شما
                                            st.markdown(
                                                f"""
                                                <div style="background-color: {res['رنگ']}22; border-left: 5px solid {res['رنگ']}; padding: 10px; margin-bottom: 5px; border-radius: 4px;">
                                                    <strong>🗓️ تاریخ:</strong> {res['تاریخ']} | 
                                                    <strong>💼 ماموریت:</strong> {res['شروع ماموریت']} تا {res['پایان ماموریت']} | 
                                                    <strong>⏰ تردد ثبت شده:</strong> {res['شروع تردد']} تا {res['پایان تردد']} <br>
                                                    <strong>📌 وضعیت:</strong> <span style="color: {res['رنگ']}; font-weight: bold;">{res['وضعیت بررسی']}</span>
                                                </div>
                                                """, 
                                                unsafe_allow_html=True
                                            )
                                    else:
                                        st.error("فایل تردد آپلود شده ساختار استاندارد ندارد. (ستون‌های 'روز'، 'شروع' یا 'پایان' یافت نشدند)")
                                except Exception as ex:
                                    st.error(f"خطا در پردازش فایل تردد: {ex}")
            
            st.markdown("---")
            st.subheader("📥 دانلود فایل خروجی")
            
            # دکمه پردازش و ساخت فایل خروجی اکسل با استایل خط قرمز (بخش قبلی کاملاً حفظ شده است)
            if st.button("🔄 پردازش و اعمال خط تیره قرمز روی ردیف‌های انتخابی"):
                if not st.session_state.selected_personnel:
                    st.warning("هیچ پرسنلی انتخاب نشده است. فایل خروجی بدون تغییر خواهد بود.")
                
                # بازنشانی نشانگر فایل آپلود شده برای خواندن مجدد توسط openpyxl
                uploaded_file.seek(0)
                wb = openpyxl.load_workbook(uploaded_file)
                ws = wb.active
                
                # پیدا کردن اندیس ستون‌های "کد پرسنلی" و "نام و نام خانوادگی" در اکسل
                header_row = [cell.value for cell in ws[1]]
                
                try:
                    id_col_idx = header_row.index('کد پرسنلی') + 1
                    name_col_idx = header_row.index('نام و نام خانوادگی') + 1
                    
                    # تعریف استایل فونت: رنگ قرمز و دارای خط خوردگی (Strikethrough)
                    red_strikethrough_font = Font(strikethrough=True, color="FF0000")
                    
                    # پیمایش ردیف‌های اکسل اصلی
                    marked_rows_count = 0
                    for row_idx in range(2, ws.max_row + 1):
                        cell_id = str(ws.cell(row=row_idx, column=id_col_idx).value).strip()
                        cell_name = str(ws.cell(row=row_idx, column=name_col_idx).value).strip()
                        
                        current_row_key = f"{cell_id} - {cell_name}"
                        
                        if current_row_key in st.session_state.selected_personnel:
                            # اعمال فونت قرمز خط‌خورده روی تمام سلول‌های ردیف منطبق
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = red_strikethrough_font
                            marked_rows_count += 1
                    
                    # ذخیره فایل در حافظه موقت (BytesIO)
                    output_stream = BytesIO()
                    wb.save(output_stream)
                    output_stream.seek(0)
                    
                    st.success(f"عملیات با موفقیت انجام شد! تعداد {marked_rows_count} ردیف در فایل نهایی خط‌خوردگی قرمز دریافت کردند.")
                    
                    # ایجاد دکمه دانلود نهایی
                    st.download_button(
                        label="📥 دانلود فایل اکسل اصلاح شده (حاوی خط خوردگی‌ها)",
                        data=output_stream,
                        file_name="processed_missions.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except ValueError:
                    st.error("خطا در ساختار هدر فایل اکسل اصلی. مطمئن شوید ستون‌های 'کد پرسنلی' و 'نام و نام خانوادگی' وجود دارند.")
                    
    except Exception as e:
        st.error(f"خطا در خواندن یا پردازش فایل: {e}")
